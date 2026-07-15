# -*- coding: utf-8 -*-
"""
=============================================================================
GERADOR DE DADOS SINTÉTICOS - DASHBOARD DE MONITORIZAÇÃO DE NITRATOS
=============================================================================
Cria um ficheiro Excel de demonstração com a mesma estrutura do ficheiro
de monitorização real:

  - Uma folha por município: Barcelos, Póvoa de Varzim, Esposende
  - Linha 3 : IDs dos poços (coluna B em diante)
  - Linhas 4+: coluna A = data (mensal), restantes colunas = NO3 (mg/L)

Os valores são FICTÍCIOS, gerados com componentes realistas:
  - nível de base próprio de cada poço
  - sazonalidade anual (pico no final do inverno, após lixiviação)
  - tendência plurianual (crescente, decrescente ou estável)
  - ruído aleatório
  - valores em falta ocasionais (5%)
  - poços que entram em monitorização mais tarde
  - linhas de datas futuras pré-preenchidas sem valores
    (reproduz o padrão típico dos ficheiros de campo)

Uso:
    python gerar_dados_sinteticos.py
=============================================================================
"""

from datetime import date

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
FICHEIRO_SAIDA = "dados_demo_zv.xlsx"
SEED = 42                      # reprodutibilidade
DATA_INICIO = date(2023, 1, 15)
N_MESES_COM_DADOS = 42         # ~3,5 anos de medições
N_MESES_FUTUROS = 4            # linhas de datas futuras sem valores

# Definição dos poços por município:
#   (id, base mg/L, tendência mg/L/ano, amplitude sazonal, mês de início [0 = desde o início])
POCOS = {
    "Barcelos": [
        ("BA1", 45, +12.0, 8, 0),    # poço problemático, tendência forte
        ("BA2", 30, +1.5, 5, 0),
        ("BA3", 38, +2.0, 6, 0),
        ("BA4", 33, -1.0, 5, 0),
        ("BA5", 68, +1.0, 7, 0),     # acima do VMA de forma crónica
        ("BA6", 8, 0.0, 2, 0),       # poço limpo
        ("BA7", 28, -3.0, 9, 0),
        ("BA8", 31, -1.5, 5, 0),
        ("BA9", 15, +0.5, 3, 6),     # entra 6 meses depois
        ("BA10", 26, 0.0, 4, 0),
        ("BA11", 42, +3.0, 6, 12),   # entra 1 ano depois
    ],
    "Póvoa de Varzim": [
        ("PV1", 52, -2.0, 8, 0),
        ("PV2", 36, +1.0, 6, 0),
        ("PV3", 24, 0.0, 4, 0),
        ("PV4", 61, +2.5, 9, 0),
        ("PV5", 18, -0.5, 3, 0),
        ("PV6", 44, +1.5, 7, 9),
        ("PV99", 55, 0.0, 6, 0),     # suspenso (excluído por omissão na app)
    ],
    "Esposende": [
        ("ESP1", 41, +2.0, 7, 0),
        ("ESP2", 29, -1.0, 5, 0),
        ("ESP3", 57, +4.0, 8, 0),
        ("ESP4", 22, 0.0, 4, 0),
        ("ESP5", 48, 0.0, 6, 0),     # suspenso
        ("ESP6", 35, 0.0, 5, 0),     # suspenso
        ("ESP7", 12, +0.5, 2, 3),
    ],
}

PCT_FALTA = 0.05   # probabilidade de valor em falta em cada medição


# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
def gerar_serie(rng, base, tendencia_ano, amp_sazonal, mes_inicio, n_meses):
    """Gera a série mensal de um poço. Devolve lista de valores (None = falta)."""
    valores = []
    for m in range(n_meses):
        if m < mes_inicio:
            valores.append(None)          # ainda não monitorizado
            continue
        anos = m / 12.0
        # Sazonalidade: pico ~fevereiro-março (lixiviação de inverno)
        mes_do_ano = (DATA_INICIO.month - 1 + m) % 12
        sazonal = amp_sazonal * np.cos(2 * np.pi * (mes_do_ano - 2) / 12)
        ruido = rng.normal(0, base * 0.06 + 1.5)
        v = base + tendencia_ano * anos + sazonal + ruido
        v = max(0.5, v)                    # nitratos nunca negativos
        if rng.random() < PCT_FALTA:
            valores.append(None)           # medição em falta
        else:
            valores.append(round(float(v), 1))
    return valores


def datas_mensais(inicio: date, n: int) -> list:
    """Gera n datas mensais (dia 15 de cada mês) a partir de 'inicio'."""
    datas = []
    ano, mes = inicio.year, inicio.month
    for _ in range(n):
        datas.append(date(ano, mes, 15))
        mes += 1
        if mes > 12:
            mes, ano = 1, ano + 1
    return datas


def main():
    rng = np.random.default_rng(SEED)
    n_total = N_MESES_COM_DADOS + N_MESES_FUTUROS
    datas = datas_mensais(DATA_INICIO, n_total)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_suspenso = PatternFill("solid", start_color="FFC7CE")
    negrito = Font(bold=True)

    for municipio, pocos in POCOS.items():
        ws = wb.create_sheet(municipio)

        # Linhas 1-2: título (como no ficheiro de campo)
        ws.cell(row=1, column=1,
                value=f"Monitorização de Nitratos — {municipio} (DADOS SINTÉTICOS)").font = negrito
        ws.cell(row=2, column=1, value="NO3 (mg/L)")

        # Linha 3: IDs dos poços (coluna B em diante)
        ws.cell(row=3, column=1, value="Data").font = negrito
        suspensos = {"PV99", "ESP5", "ESP6"}
        for j, (pid, *_rest) in enumerate(pocos, start=2):
            c = ws.cell(row=3, column=j, value=pid)
            c.font = negrito
            if pid in suspensos:
                c.fill = fill_suspenso     # cor = suspenso (convenção do ficheiro real)

        # Séries de cada poço
        series = {}
        for pid, base, tend, amp, inicio in pocos:
            vals = gerar_serie(rng, base, tend, amp, inicio, N_MESES_COM_DADOS)
            vals += [None] * N_MESES_FUTUROS   # datas futuras sem valores
            series[pid] = vals

        # Linhas 4+: datas e valores
        for i, d in enumerate(datas):
            r = 4 + i
            ws.cell(row=r, column=1, value=d)
            ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
            for j, (pid, *_rest) in enumerate(pocos, start=2):
                v = series[pid][i]
                if v is not None:
                    ws.cell(row=r, column=j, value=v)

        ws.column_dimensions["A"].width = 12

    wb.save(FICHEIRO_SAIDA)

    n_pocos = sum(len(p) for p in POCOS.values())
    print(f"[OK] '{FICHEIRO_SAIDA}' gerado:")
    print(f"     {len(POCOS)} municípios, {n_pocos} poços, "
          f"{N_MESES_COM_DADOS} meses com dados + {N_MESES_FUTUROS} datas futuras vazias")
    print(f"     Período: {datas[0]} a {datas[N_MESES_COM_DADOS-1]} (dados) "
          f"/ até {datas[-1]} (datas pré-preenchidas)")
    print("     ATENÇÃO: valores fictícios, apenas para demonstração.")


if __name__ == "__main__":
    main()
