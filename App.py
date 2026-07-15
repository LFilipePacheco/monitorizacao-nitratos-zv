# -*- coding: utf-8 -*-
"""
=============================================================================
DASHBOARD DE MONITORIZAÇÃO DE NITRATOS
Zona Vulnerável Esposende – Vila do Conde | CCDR-Norte, IP
=============================================================================
Visualização dinâmica dos dados de monitorização de poços agrícolas.

Execução local:
    pip install -r requirements.txt
    streamlit run app.py

Estrutura esperada do Excel (uma folha por município):
    - Linha 3  : IDs dos poços (coluna B em diante)
    - Linhas 4+: coluna A = data, restantes colunas = valores NO3 (mg/L)
=============================================================================
"""

import io
import os
from datetime import datetime, date

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy import stats as sp_stats

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Nitratos ZV Esposende–Vila do Conde",
    page_icon="💧",
    layout="wide",
)

# ─────────────────────────────────────────────────────────────────────────────
# GATE DE AUTENTICAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
# A password é lida de st.secrets["password"]:
#   - Localmente : ficheiro .streamlit/secrets.toml  →  password = "..."
#   - Streamlit Cloud / HF Spaces: definir o secret no painel da app
# Se não existir nenhum secret configurado, a app corre SEM autenticação
# (útil em desenvolvimento local).

def _gate_autenticacao() -> bool:
    """Devolve True se o utilizador estiver autenticado."""
    import hmac

    # Sem secret configurado → sem gate (modo desenvolvimento)
    try:
        password_correta = st.secrets["password"]
    except (KeyError, FileNotFoundError):
        return True

    if st.session_state.get("autenticado", False):
        return True

    st.title("🔒 Acesso restrito")
    st.markdown(
        "**Monitorização de Nitratos — ZV Esposende–Vila do Conde**  \n"
        "CCDR-Norte, IP · Divisão Agroalimentar e Pescas")

    with st.form("login"):
        password = st.text_input("Password", type="password")
        submeter = st.form_submit_button("Entrar")

    if submeter:
        # Comparação em tempo constante (evita timing attacks)
        if hmac.compare_digest(password, password_correta):
            st.session_state["autenticado"] = True
            st.rerun()
        else:
            st.error("Password incorreta.")

    return False


if not _gate_autenticacao():
    st.stop()

# Versão de demonstração pública — dados sintéticos incluídos no repositório
EXCEL_PATH_LOCAL = ""   # (a versão de produção lê o Excel institucional)

# Nome do Excel no repositório (deployment na Streamlit Cloud)
# ATENÇÃO: tem de coincidir EXACTAMENTE com o nome do ficheiro no GitHub
EXCEL_PATH_REPO = "dados_demo_zv.xlsx"

FOLHAS_MUNICIPIOS = ["Barcelos", "Póvoa de Varzim", "Esposende"]

# Poços suspensos: detectados AUTOMATICAMENTE pela cor de preenchimento
# da célula do ID na linha 3 do Excel (convenção do ficheiro de campo).
# Esta lista serve apenas de RESERVA caso a detecção não encontre nenhum.
POCOS_SUSPENSOS_FALLBACK = {
    "Barcelos": [],
    "Póvoa de Varzim": ["PV99"],
    "Esposende": ["ESP5", "ESP6"],
}

# Filtro de sanidade: valores fora deste intervalo são ignorados na leitura.
# Tecto a 1000 (não 500): existem medições reais acima de 500 mg/L na série
# histórica (PV3 em 2008: 501-591; PV7 em 2026: 600). O filtro serve para
# rejeitar disparates óbvios (coordenadas GPS, datas serializadas), não
# contaminação extrema genuína.
VALOR_MIN, VALOR_MAX = 0.0, 1000.0

CORES_MUNICIPIO = {
    "Barcelos":        "#1f77b4",
    "Póvoa de Varzim": "#ff7f0e",
    "Esposende":       "#2ca02c",
}

# ─────────────────────────────────────────────────────────────────────────────
# LEITURA E PRÉ-PROCESSAMENTO (mesma lógica do script batch)
# ─────────────────────────────────────────────────────────────────────────────
def _tem_preenchimento(cell) -> bool:
    """True se a célula tiver preenchimento sólido não-branco
    (convenção do Excel de campo para poços suspensos)."""
    try:
        fill = cell.fill
        if fill is None or fill.patternType != "solid":
            return False
        rgb = getattr(fill.start_color, "rgb", None)
        # Sem cor definida, preto-transparente ou branco → não conta
        return rgb not in (None, "00000000", "FFFFFFFF")
    except Exception:
        return False


@st.cache_data(show_spinner="A ler o ficheiro Excel...")
def ler_folha(conteudo_bytes: bytes, sheet_name: str):
    """Lê uma folha e devolve (DataFrame, lista de poços suspensos,
    n.º de valores fora do intervalo de sanidade).
    Suspensos = células da linha 3 com preenchimento de cor."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame(), [], 0
    ws = wb[sheet_name]

    # Linha 3 → IDs dos poços (+ detecção de suspensos pela cor)
    ids_poco = {}
    suspensos = []
    for cell in ws[3]:
        if cell.column == 1:
            continue
        if cell.value and isinstance(cell.value, str):
            poco_id = cell.value.strip()
            ids_poco[cell.column] = poco_id
            if _tem_preenchimento(cell):
                suspensos.append(poco_id)

    registos = []
    n_invalidos = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        data_val = row[0]
        if not isinstance(data_val, datetime):
            continue
        entry = {"data": data_val.date()}
        for col_idx, poco_id in ids_poco.items():
            val = row[col_idx - 1]
            if isinstance(val, (int, float)) and not np.isnan(float(val)):
                v = float(val)
                if VALOR_MIN <= v <= VALOR_MAX:
                    entry[poco_id] = v
                else:
                    entry[poco_id] = np.nan
                    n_invalidos += 1
            else:
                entry[poco_id] = np.nan
        registos.append(entry)

    if not registos:
        return pd.DataFrame(), suspensos, n_invalidos
    df = pd.DataFrame(registos).set_index("data")
    df.index = pd.to_datetime(df.index)
    return df.sort_index(), suspensos, n_invalidos


@st.cache_data(show_spinner=False)
def carregar_dados(conteudo_bytes: bytes):
    """Carrega todas as folhas.
    Devolve (dados, suspensos_detectados, n_valores_invalidos):
      dados                {municipio: DataFrame} com todos os poços
      suspensos_detectados {municipio: [ids]} pela cor da linha 3
      n_valores_invalidos  total de valores fora de [VALOR_MIN, VALOR_MAX]
    """
    dados, suspensos_det, invalidos = {}, {}, 0
    for mun in FOLHAS_MUNICIPIOS:
        df, susp, n_inv = ler_folha(conteudo_bytes, mun)
        invalidos += n_inv
        if not df.empty:
            dados[mun] = df
            # Detecção pela cor tem prioridade; fallback só se nada detectado
            suspensos_det[mun] = susp if susp else POCOS_SUSPENSOS_FALLBACK.get(mun, [])
    return dados, suspensos_det, invalidos


def imputar(df: pd.DataFrame, metodo: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Imputa nulos APENAS em lacunas interiores do período activo de cada
    poço — entre a primeira e a última observação real. Datas futuras já
    pré-preenchidas no Excel (sem medição) nunca são imputadas.
    Devolve (df_imputado, tabela_qualidade)."""
    df_out = df.copy()
    linhas = []
    for poco in df.columns:
        serie = df[poco]
        first_idx = serie.first_valid_index()
        last_idx  = serie.last_valid_index()
        if first_idx is None:
            linhas.append({"Poço": poco, "Total obs.": 0, "Em falta": 0,
                           "% em falta": 0.0, "Valor imputado": "—"})
            continue
        # Período activo = da 1.ª à ÚLTIMA observação real (lacunas interiores)
        activa = serie[first_idx:last_idx]
        em_falta = int(activa.isna().sum())
        total = len(activa)
        valor_imp = None
        if em_falta > 0:
            valid = serie.dropna()
            if len(valid) >= 3:   # mínimo de obs. reais para imputação fiável
                valor_imp = valid.median() if metodo == "mediana" else valid.mean()
                mask = ((df_out.index >= first_idx) & (df_out.index <= last_idx)
                        & df_out[poco].isna())
                df_out.loc[mask, poco] = valor_imp
        linhas.append({
            "Poço": poco,
            "Total obs.": total,
            "Em falta": em_falta,
            "% em falta": round(100 * em_falta / total, 1) if total else 0.0,
            # Sempre string (tipo único na coluna — exigência do PyArrow/st.dataframe)
            "Valor imputado": f"{valor_imp:.2f}" if valor_imp is not None else "—",
        })
    return df_out, pd.DataFrame(linhas)


def estatisticas_poco(df_orig: pd.DataFrame, vma: float, alerta: float) -> pd.DataFrame:
    """Estatísticas descritivas sobre observações reais (sem imputação)."""
    linhas = []
    for poco in df_orig.columns:
        s = df_orig[poco].dropna()
        if s.empty:
            continue
        linhas.append({
            "Poço": poco,
            "N": len(s),
            "Média": round(s.mean(), 1),
            "Mediana": round(s.median(), 1),
            "Desvio P.": round(s.std(), 1),
            "Mín.": round(s.min(), 1),
            "Máx.": round(s.max(), 1),
            "CV (%)": round(s.std() / s.mean() * 100, 1) if s.mean() else np.nan,
            "> Alerta": int((s > alerta).sum()),
            "> VMA": int((s > vma).sum()),
            "Conform. (%)": round((s <= vma).mean() * 100, 1),
        })
    return pd.DataFrame(linhas)


def linha_limiares(fig, vma, alerta):
    """Adiciona linhas horizontais de VMA e alerta a uma figura Plotly."""
    fig.add_hline(y=vma, line_dash="dash", line_color="red",
                  annotation_text=f"VMA {vma} mg/L",
                  annotation_position="top left")
    fig.add_hline(y=alerta, line_dash="dot", line_color="orange",
                  annotation_text=f"Alerta {alerta} mg/L",
                  annotation_position="bottom left")
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# BARRA LATERAL — FONTE DE DADOS E FILTROS
# ─────────────────────────────────────────────────────────────────────────────
st.sidebar.title("💧 Nitratos ZV")
st.sidebar.caption("Zona Vulnerável Esposende – Vila do Conde")
st.sidebar.info("🧪 DEMONSTRAÇÃO — dados sintéticos (valores fictícios)")

if st.session_state.get("autenticado", False):
    if st.sidebar.button("🚪 Terminar sessão"):
        st.session_state["autenticado"] = False
        st.rerun()

ficheiro = st.sidebar.file_uploader(
    "Ficheiro Excel de monitorização", type=["xlsx"],
    help="Estrutura: linha 3 = IDs dos poços; linhas 4+ = datas e valores.")

conteudo = None
if ficheiro is not None:
    conteudo = ficheiro.getvalue()
elif os.path.exists(EXCEL_PATH_LOCAL):
    with open(EXCEL_PATH_LOCAL, "rb") as f:
        conteudo = f.read()
    st.sidebar.info("A usar o ficheiro local por omissão.")
else:
    # Excel do repositório: tentar o nome exacto e, se falhar (variações
    # Unicode nos acentos do nome), qualquer .xlsx na raiz do repositório
    import glob
    candidatos = ([EXCEL_PATH_REPO] if os.path.exists(EXCEL_PATH_REPO)
                  else sorted(glob.glob("*.xlsx")))
    if candidatos:
        caminho_repo = candidatos[0]
        with open(caminho_repo, "rb") as f:
            conteudo = f.read()
        mtime = datetime.fromtimestamp(os.path.getmtime(caminho_repo))
        st.sidebar.caption(f"Dados do repositório · actualizados em {mtime:%d/%m/%Y}")

if conteudo is None:
    # ── Ecrã de boas-vindas com instruções ───────────────────────────────────
    st.title("💧 Monitorização de Nitratos")
    st.subheader("Zona Vulnerável Esposende – Vila do Conde")
    st.caption("CCDR-Norte, IP · Divisão Agroalimentar e Pescas")
    st.divider()

    st.markdown("### Como começar")
    st.markdown(
        "Esta aplicação analisa o ficheiro Excel de monitorização dos poços "
        "(**Monitorizaçao_Poços_ZV.xlsx**) — o mesmo onde regista as medições "
        "mensais. Os dados são processados apenas na sua sessão e **não ficam "
        "guardados em nenhum servidor**.")

    st.markdown("""
**Passo 1 — Localize o ficheiro no seu computador**
Abra a pasta onde guarda o Excel de monitorização (normalmente no OneDrive,
em *Zona Vulneravel → Dados → Poços*). Se o ficheiro estiver aberto no Excel,
guarde e feche-o primeiro.

**Passo 2 — Carregue o ficheiro nesta página**
Use a caixa abaixo (ou a da barra lateral, à esquerda): clique em
**"Browse files"** e escolha o ficheiro, ou simplesmente **arraste-o**
para cima da caixa.

**Passo 3 — Explore os resultados**
A análise abre automaticamente. Na barra lateral pode ajustar o período,
os municípios e os poços a analisar. Os gráficos são interactivos:
passe o rato para ver valores, e use os ícones no canto de cada gráfico
para ampliar ou descarregar como imagem.
""")

    ficheiro_central = st.file_uploader(
        "📂 Carregar o ficheiro Excel de monitorização",
        type=["xlsx"],
        key="uploader_central",
        help="Ficheiro .xlsx com as folhas Barcelos, Póvoa de Varzim e Esposende "
             "(linha 3 = IDs dos poços; linhas seguintes = datas e valores).")

    if ficheiro_central is not None:
        conteudo = ficheiro_central.getvalue()
    else:
        with st.expander("❓ Problemas frequentes"):
            st.markdown("""
- **O botão não faz nada** — verifique se o ficheiro tem a extensão `.xlsx`
  (a caixa só aceita este formato).
- **"Não foi possível ler dados"** — confirme que o ficheiro é o de
  monitorização dos poços, com as folhas *Barcelos*, *Póvoa de Varzim*
  e *Esposende*, e os IDs dos poços na linha 3.
- **Os dados parecem desatualizados** — certifique-se de que guardou o
  Excel antes de o carregar aqui; a app lê o ficheiro tal como está
  no disco no momento do carregamento.
- **Registei novas medições** — basta guardar o Excel e voltar a
  carregá-lo aqui; a análise actualiza de imediato.
""")
        st.stop()

dados_raw, suspensos_detectados, n_invalidos = carregar_dados(conteudo)
if not dados_raw:
    st.error("Não foi possível ler dados de nenhuma folha (Barcelos, Póvoa de Varzim, Esposende).")
    st.stop()

if n_invalidos > 0:
    st.sidebar.warning(
        f"⚠️ {n_invalidos} valor(es) fora do intervalo "
        f"{VALOR_MIN:.0f}–{VALOR_MAX:.0f} mg/L foram ignorados na leitura. "
        "Verifique o Excel de origem.")

# Filtros
st.sidebar.divider()
st.sidebar.subheader("Filtros")

municipios_sel = st.sidebar.multiselect(
    "Municípios", options=list(dados_raw.keys()),
    default=list(dados_raw.keys()))

data_min_global = min(df.index.min() for df in dados_raw.values()).date()
data_max_global = max(df.index.max() for df in dados_raw.values()).date()

periodo = st.sidebar.date_input(
    "Período de análise",
    value=(date(2026, 1, 1) if date(2026, 1, 1) >= data_min_global else data_min_global,
           data_max_global),
    min_value=data_min_global, max_value=data_max_global)
if isinstance(periodo, tuple) and len(periodo) == 2:
    dt_ini, dt_fim = pd.to_datetime(periodo[0]), pd.to_datetime(periodo[1])
else:
    dt_ini, dt_fim = pd.to_datetime(periodo), pd.to_datetime(data_max_global)

metodo_imp = st.sidebar.radio("Imputação de dados em falta",
                              ["mediana", "media", "nenhuma"], horizontal=True)

st.sidebar.divider()
st.sidebar.subheader("Limiares (mg NO₃/L)")
vma = st.sidebar.number_input("VMA", value=50, min_value=1, step=5)
alerta = st.sidebar.number_input("Limiar de alerta", value=25, min_value=1, step=5)

st.sidebar.divider()
st.sidebar.subheader("Poços suspensos (excluídos)")
st.sidebar.caption("Pré-seleccionados pela cor de preenchimento na linha 3 do Excel.")
suspensos_sel = {}
for mun in municipios_sel:
    todos_pocos = list(dados_raw[mun].columns)
    suspensos_sel[mun] = st.sidebar.multiselect(
        f"{mun}", options=todos_pocos,
        default=[p for p in suspensos_detectados.get(mun, []) if p in todos_pocos],
        key=f"susp_{mun}")

# ─────────────────────────────────────────────────────────────────────────────
# PREPARAR DADOS FILTRADOS
# ─────────────────────────────────────────────────────────────────────────────
dados_filt = {}     # observações reais (para estatísticas/conformidade)
dados_imp = {}      # com imputação (para gráficos de séries)
qualidade = {}

for mun in municipios_sel:
    df = dados_raw[mun].drop(columns=suspensos_sel.get(mun, []), errors="ignore")
    df = df[(df.index >= dt_ini) & (df.index <= dt_fim)].dropna(axis=1, how="all")
    if df.empty or df.shape[1] == 0:
        continue
    dados_filt[mun] = df
    if metodo_imp in ("mediana", "media"):
        df_i, qual = imputar(df, metodo_imp)
    else:
        df_i, qual = df.copy(), imputar(df, "mediana")[1].assign(**{"Valor imputado": "—"})
    dados_imp[mun] = df_i
    qualidade[mun] = qual

if not dados_filt:
    st.warning("Sem dados no período/filtros seleccionados.")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# CABEÇALHO E KPIs GLOBAIS
# ─────────────────────────────────────────────────────────────────────────────
st.title("Monitorização de Nitratos")
st.markdown(
    f"**Zona Vulnerável Esposende – Vila do Conde** · CCDR-Norte, IP — "
    f"Divisão Agroalimentar e Pescas  \n"
    f"Período: {dt_ini.strftime('%d/%m/%Y')} → {dt_fim.strftime('%d/%m/%Y')} · "
    f"Imputação: {metodo_imp}")

todos_vals = pd.concat([df.stack() for df in dados_filt.values()])
n_pocos = sum(df.shape[1] for df in dados_filt.values())
n_exced = int((todos_vals > vma).sum())
pct_conf = round((todos_vals <= vma).mean() * 100, 1)

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Poços activos", n_pocos)
c2.metric("Observações", len(todos_vals))
c3.metric("Média global", f"{todos_vals.mean():.1f} mg/L")
c4.metric("Ultrapassagens VMA", n_exced)
c5.metric("Conformidade", f"{pct_conf} %",
          delta=None, help=f"Observações ≤ {vma} mg/L")

# ─────────────────────────────────────────────────────────────────────────────
# SEPARADORES
# ─────────────────────────────────────────────────────────────────────────────
tab_series, tab_dist, tab_heat, tab_tend, tab_conf, tab_qual, tab_dados = st.tabs(
    ["📈 Séries temporais", "📦 Distribuição", "🔥 Heatmap anual",
     "📉 Tendências", "✅ Conformidade", "🧪 Qualidade dos dados", "🗃️ Dados"])

# ── Séries temporais ─────────────────────────────────────────────────────────
with tab_series:
    mun_sel = st.selectbox("Município", options=list(dados_filt.keys()), key="mun_series")
    df_real = dados_filt[mun_sel]      # apenas observações medidas
    df_com_imp = dados_imp[mun_sel]    # com lacunas interiores imputadas
    pocos_sel = st.multiselect("Poços", options=list(df_real.columns),
                               default=list(df_real.columns), key="pocos_series")
    mostrar_imp = st.toggle(
        "Incluir valores imputados (marcadores abertos ◇)",
        value=False,
        help="Por omissão, o gráfico mostra apenas as medições reais. "
             "Os valores imputados preenchem lacunas interiores e são "
             "assinalados com marcadores abertos.")
    if pocos_sel:
        base = df_com_imp if mostrar_imp else df_real
        df_long = (base[pocos_sel].reset_index()
                   .melt(id_vars="data", var_name="Poço", value_name="NO₃ (mg/L)")
                   .dropna())
        fig = px.line(df_long, x="data", y="NO₃ (mg/L)", color="Poço",
                      markers=True,
                      title=f"Evolução temporal — {mun_sel}"
                            + (" (com imputação)" if mostrar_imp else " (medições reais)"))
        linha_limiares(fig, vma, alerta)

        # Sobrepor os pontos imputados com marcador aberto
        if mostrar_imp:
            mask_imp = df_real[pocos_sel].isna() & df_com_imp[pocos_sel].notna()
            df_imp_pts = (df_com_imp[pocos_sel].where(mask_imp)
                          .reset_index()
                          .melt(id_vars="data", var_name="Poço",
                                value_name="NO₃ (mg/L)").dropna())
            if not df_imp_pts.empty:
                fig.add_trace(go.Scatter(
                    x=df_imp_pts["data"], y=df_imp_pts["NO₃ (mg/L)"],
                    mode="markers",
                    marker=dict(symbol="diamond-open", size=10,
                                color="black", line=dict(width=1.5)),
                    name="Imputado"))

        fig.update_layout(height=520, hovermode="x unified",
                          xaxis_title="Data")
        st.plotly_chart(fig, use_container_width=True)

        # Médias mensais
        with st.expander("Médias mensais (sazonalidade)"):
            df_mes = df_real[pocos_sel].resample("ME").mean()
            df_mes_long = (df_mes.reset_index()
                           .melt(id_vars="data", var_name="Poço",
                                 value_name="NO₃ (mg/L)").dropna())
            fig_m = px.line(df_mes_long, x="data", y="NO₃ (mg/L)",
                            color="Poço", markers=True,
                            title=f"Médias mensais — {mun_sel} (medições reais)")
            linha_limiares(fig_m, vma, alerta)
            fig_m.update_layout(height=420)
            st.plotly_chart(fig_m, use_container_width=True)

# ── Distribuição ─────────────────────────────────────────────────────────────
with tab_dist:
    col_a, col_b = st.columns(2)

    with col_a:
        mun_box = st.selectbox("Município (boxplot)",
                               options=list(dados_filt.keys()), key="mun_box")
        df_b = dados_filt[mun_box]
        df_b_long = (df_b.reset_index()
                     .melt(id_vars="data", var_name="Poço",
                           value_name="NO₃ (mg/L)").dropna())
        fig_box = px.box(df_b_long, x="Poço", y="NO₃ (mg/L)", points="outliers",
                         color_discrete_sequence=[CORES_MUNICIPIO.get(mun_box, "#5599cc")],
                         title=f"Distribuição por poço — {mun_box}")
        linha_limiares(fig_box, vma, alerta)
        fig_box.update_layout(height=480)
        st.plotly_chart(fig_box, use_container_width=True)

    with col_b:
        registos = []
        for mun, df in dados_filt.items():
            vals = df.stack().dropna()
            registos.append(pd.DataFrame(
                {"Município": mun, "NO₃ (mg/L)": vals.values}))
        df_all = pd.concat(registos, ignore_index=True)
        fig_v = px.violin(df_all, x="Município", y="NO₃ (mg/L)", box=True,
                          color="Município",
                          color_discrete_map=CORES_MUNICIPIO,
                          title="Comparação entre municípios")
        linha_limiares(fig_v, vma, alerta)
        fig_v.update_layout(height=480, showlegend=False)
        st.plotly_chart(fig_v, use_container_width=True)

# ── Heatmap anual ────────────────────────────────────────────────────────────
with tab_heat:
    mun_h = st.selectbox("Município", options=list(dados_filt.keys()), key="mun_heat")
    df_anual = dados_filt[mun_h].resample("YE").mean().round(1)
    df_anual.index = df_anual.index.year
    if df_anual.shape[0] < 1:
        st.info("Sem dados suficientes para o heatmap anual.")
    else:
        fig_h = px.imshow(df_anual, text_auto=".0f", aspect="auto",
                          color_continuous_scale="YlOrRd",
                          labels=dict(x="Poço", y="Ano", color="NO₃ (mg/L)"),
                          title=f"Média anual de nitratos — {mun_h}")
        fig_h.update_layout(height=380 + 30 * len(df_anual))
        st.plotly_chart(fig_h, use_container_width=True)

# ── Tendências ───────────────────────────────────────────────────────────────
with tab_tend:
    mun_t = st.selectbox("Município", options=list(dados_filt.keys()), key="mun_tend")
    df_t = dados_filt[mun_t]
    linhas_t = []
    fig_t = go.Figure()
    palette = px.colors.qualitative.T10
    for i, poco in enumerate(df_t.columns):
        s = df_t[poco].dropna()
        if len(s) < 3:
            continue
        x_num = (s.index - s.index[0]).days.values.astype(float)
        slope, intercept, r, p, _ = sp_stats.linregress(x_num, s.values)
        cor = palette[i % len(palette)]
        fig_t.add_trace(go.Scatter(x=s.index, y=s.values, mode="markers",
                                   marker=dict(size=4, color=cor, opacity=0.45),
                                   name=poco, legendgroup=poco, showlegend=False))
        y_line = intercept + slope * np.array([x_num[0], x_num[-1]])
        fig_t.add_trace(go.Scatter(x=[s.index[0], s.index[-1]], y=y_line,
                                   mode="lines", line=dict(width=3, color=cor),
                                   name=f"{poco} ({'↑' if slope > 0 else '↓'})",
                                   legendgroup=poco))
        linhas_t.append({
            "Poço": poco,
            "Declive (mg/L/ano)": round(slope * 365.25, 2),
            "R²": round(r ** 2, 3),
            "p-value": round(p, 4),
            "Tendência": ("Crescente" if slope > 0 else "Decrescente")
                         + (" *" if p < 0.05 else " (n.s.)"),
        })
    linha_limiares(fig_t, vma, alerta)
    fig_t.update_layout(height=520, title=f"Tendência linear por poço — {mun_t}",
                        yaxis_title="NO₃ (mg/L)")
    st.plotly_chart(fig_t, use_container_width=True)
    if linhas_t:
        st.caption("* estatisticamente significativa (p < 0,05); n.s. = não significativa")
        st.dataframe(pd.DataFrame(linhas_t), use_container_width=True, hide_index=True)

# ── Conformidade ─────────────────────────────────────────────────────────────
with tab_conf:
    registos_c = []
    for mun, df in dados_filt.items():
        for poco in df.columns:
            s = df[poco].dropna()
            if s.empty:
                continue
            registos_c.append({
                "Município": mun, "Poço": poco,
                "N": len(s),
                "Ultrapassagens": int((s > vma).sum()),
                "Conformidade (%)": round((s <= vma).mean() * 100, 1),
            })
    df_c = pd.DataFrame(registos_c)
    fig_c = px.bar(df_c, x="Poço", y="Conformidade (%)", color="Município",
                   color_discrete_map=CORES_MUNICIPIO, text="Conformidade (%)",
                   title=f"Conformidade com o VMA (≤ {vma} mg/L) por poço")
    fig_c.add_hline(y=100, line_dash="dash", line_color="green")
    fig_c.add_hline(y=80, line_dash="dot", line_color="orange",
                    annotation_text="80 % referência")
    fig_c.update_traces(texttemplate="%{text:.0f}%", textposition="outside")
    fig_c.update_layout(height=500, yaxis_range=[0, 112])
    st.plotly_chart(fig_c, use_container_width=True)

    st.subheader("Estatísticas descritivas por poço")
    for mun in dados_filt:
        st.markdown(f"**{mun}**")
        st.dataframe(estatisticas_poco(dados_filt[mun], vma, alerta),
                     use_container_width=True, hide_index=True)

# ── Qualidade dos dados ──────────────────────────────────────────────────────
with tab_qual:
    st.markdown(
        "Contagem de valores em falta **dentro do período activo** de cada poço "
        "(a partir da primeira observação real). "
        f"Método de imputação seleccionado: **{metodo_imp}**.")
    for mun, qual in qualidade.items():
        st.markdown(f"**{mun}**")
        st.dataframe(qual, use_container_width=True, hide_index=True)

# ── Dados ────────────────────────────────────────────────────────────────────
with tab_dados:
    mun_d = st.selectbox("Município", options=list(dados_filt.keys()), key="mun_dados")
    ver_imp = st.toggle("Mostrar valores imputados", value=False)
    df_show = (dados_imp if ver_imp else dados_filt)[mun_d]
    st.dataframe(df_show, use_container_width=True)
    csv = df_show.to_csv(sep=";", decimal=",").encode("utf-8-sig")
    st.download_button(
        "⬇️ Descarregar CSV",
        data=csv,
        file_name=f"nitratos_{mun_d.replace(' ', '_')}_{date.today():%Y%m%d}.csv",
        mime="text/csv")

st.divider()
st.caption(
    f"Fonte: monitorização de poços agrícolas · Diretiva Nitratos (91/676/CEE) · "
    f"VMA = {vma} mg NO₃/L · Gerado em {datetime.now():%d/%m/%Y %H:%M}")
